import openpyxl

wb = openpyxl.load_workbook('imagedata.xlsx')
sheet = wb['Sheet1']


# create a set of values in few columns

csv_file = open('labels.csv', 'w')

categories = set()
for i in range(1, sheet.max_row+1):
    for j in range(3,9):
        value = sheet.cell(row=i, column=j).value
        # seprate by ,
        if value is None:
            continue
        for v in value.split(','):
            # print(v.strip())
            categories.add(v.strip())

csv_file.write('[')
for c in categories:
    csv_file.write('"'+c + '",')
# remove the last comma
csv_file.seek(csv_file.tell() - 1, 0)
csv_file.truncate()
csv_file.write(']')
csv_file.close()
wb.close()

print(categories)