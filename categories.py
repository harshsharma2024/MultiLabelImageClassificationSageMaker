import openpyxl

wb = openpyxl.load_workbook('imagedata.xlsx')
sheet = wb['Sheet1']

# create a .csv file
csv_file = open('prepare_data/img_clf_multilabel_lst/all_images_gt/clf_labels.csv', 'w')
max_row = sheet.max_row
for i in range(1,max_row+1):
        
    if sheet.cell(row=i,column=9).value=='Y':
        # store from column 3 to column 8 in a list
        # write the image name to the .csv file without quotes with the extension of .jpg, the name already containd some old extension
        name = sheet.cell(row=i, column=1).value.split('.')[0]
        csv_file.write((name + '.jpg,'))


        csv_file.write('"[')
        for j in range(3,9):
            # categories.append(sheet.cell(row=i, column=j).value)
            if sheet.cell(row=i, column=j).value != None:
                csv_file.write(sheet.cell(row=i, column=j).value + ',')
        # print(categories)
        # remove the last comma
        csv_file.seek(csv_file.tell() - 1, 0)
        csv_file.truncate()
        csv_file.write(']"')

        csv_file.write('\n')

csv_file.close()
wb.close()


