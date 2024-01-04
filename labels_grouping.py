# Open a imagedata.xlsx and form set of labels

import openpyxl
import os

wb = openpyxl.load_workbook('imagedata.xlsx')
sheet = wb['Sheet1']

# create a .csv file
csv_file = open('grouped_labels.csv', 'w')

max_row = sheet.max_row

# store all labels in a set corresponding to each cloumn
labels = []

for j in range(3,9):
    b = set()
    for i in range(1,max_row+1):
        if (sheet.cell(row=i, column=j).value != None) & (str(sheet.cell(row=i, column=j).value) != 'none') & (str(sheet.cell(row=i, column=j).value) != 'other'):
            # store its value by removing quotes and converting to lower case
            val = str(sheet.cell(row=i, column=j).value).lower().split(',')
            for p in val:
                # remove spaces
                p = p.strip()
                b.add(p)
    labels.append(b)

    csv_file.write(str(b) + '\n')