import openpyxl
import re
import requests
from urllib.parse import unquote
from PIL import Image
from io import BytesIO


# Open the workbook and select the first worksheet
wb = openpyxl.load_workbook('imagedata.xlsx')
sheet = wb['Sheet1']

# Get the max row count
max_row = sheet.max_row

# Iterate through all the rows in the excel file
for i in range(4, max_row + 1):
    # Get the cell value in the second column
    cell_obj = sheet.cell(row=i, column=2)

    # Get the original link from the cell value
    original_link = cell_obj.value

    # Extract the link from the cell value
    url_pattern = r'\((https?://[^\s]+)\)'
    match = re.search(url_pattern, original_link)
    if match:
        url = match.group(1)
        url = unquote(url)
        # Update the cell with the extracted link
        cell_obj.value = url

    url = cell_obj.value
    # Downloading the image
    image_name = sheet.cell(row=i, column=1).value
    response = requests.get(url)
    # Save the image in the folder
    image_path = 'prepare_data/img_clf_multilabel_lst/all_images_gt/' + image_name
    if response.status_code == 200:
            # Get the content of the response
            image_data = response.content
            
            # Save the image to a file
            with open(image_path, 'wb') as file:
                file.write(image_data)
                
            print(f'Image "{image_name}" downloaded successfully.')

            try:
                img = Image.open(BytesIO(image_data))
                img.show()
            except Exception as e:
                print(f'Error opening image "{image_name}": {e}')
    else:
        print(f'Failed to download image for row {i}. Status code: {response.status_code}')
# Save the excel file
wb.save('imagedata.xlsx')

# Close the excel file
wb.close()
